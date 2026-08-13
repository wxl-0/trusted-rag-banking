import re
from pathlib import Path
from typing import Any, List

import openpyxl
from openpyxl.utils import get_column_letter

from src.parser.base import Chunk


class ExcelParser:
    def __init__(self, doc_id: str, source_title: str, issuer: str,
                 publish_date: str, source_url: str, local_path: str):
        self.doc_id = doc_id
        self.source_title = source_title
        self.issuer = issuer
        self.publish_date = publish_date
        self.source_url = source_url
        self.local_path = local_path

    def parse(self) -> List[Chunk]:
        path = Path(self.local_path)
        sheets = self._read_xls(path) if path.suffix.lower() == ".xls" else self._read_xlsx(path)
        chunks: List[Chunk] = []
        for sheet_name, rows in sheets:
            chunks.extend(self._parse_sheet(sheet_name, rows))
        return chunks

    def _read_xlsx(self, path: Path) -> list[tuple[str, list[list[Any]]]]:
        workbook = openpyxl.load_workbook(path, data_only=True)
        return [
            (sheet.title, [list(row) for row in sheet.iter_rows(values_only=True)])
            for sheet in workbook.worksheets
        ]

    def _read_xls(self, path: Path) -> list[tuple[str, list[list[Any]]]]:
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("Parsing .xls files requires xlrd. Install xlrd==2.0.1.") from exc

        workbook = xlrd.open_workbook(str(path))
        return [
            (sheet.name, [sheet.row_values(row_idx) for row_idx in range(sheet.nrows)])
            for sheet in workbook.sheets()
        ]

    def _parse_sheet(self, sheet_name: str, rows: list[list[Any]]) -> list[Chunk]:
        rows = self._normalize_rows(rows)
        header_indices = self._detect_header_rows(rows)
        if not header_indices:
            return []

        unit = self._detect_unit(rows)
        period = self._detect_period([self.source_title, Path(self.local_path).name, sheet_name])
        chunks: List[Chunk] = []
        sections = [
            self._section_before_header(
                rows,
                header_idx,
                header_indices[index - 1] + 1 if index else 0,
                sheet_name,
            )
            for index, header_idx in enumerate(header_indices)
        ]
        section_rows = {row_idx for _, row_idx in sections if row_idx is not None}

        for index, header_idx in enumerate(header_indices):
            end_idx = (
                header_indices[index + 1]
                if index + 1 < len(header_indices)
                else len(rows)
            )
            header = rows[header_idx]
            data_columns = self._detect_data_columns(
                header, rows[header_idx + 1:end_idx]
            )
            if not data_columns:
                continue
            label_columns = list(range(min(data_columns)))
            base_section = [sections[index][0]] if sections[index][0] else []
            group_context: dict[int, str] = {}

            for row_idx in range(header_idx + 1, end_idx):
                row = rows[row_idx]
                if not self._has_any_value(row):
                    continue
                non_empty_labels = [
                    col_idx for col_idx in label_columns
                    if not self._is_empty(row[col_idx])
                ]
                if not non_empty_labels:
                    continue
                row_label_col = non_empty_labels[-1]
                row_label = self._cell_to_text(row[row_label_col])
                if self._is_note_row(row_label):
                    continue
                for col_idx in label_columns:
                    if col_idx >= row_label_col:
                        break
                    text = self._cell_to_text(row[col_idx])
                    if text:
                        group_context[col_idx] = text
                section_path = self._dedupe_texts([
                    *base_section,
                    *(group_context.get(col_idx, "") for col_idx in label_columns
                      if col_idx < row_label_col),
                ])

                for col_idx in data_columns:
                    value = row[col_idx]
                    if self._is_empty(value) or not self._looks_like_data_cell(value):
                        continue
                    column_header = self._cell_to_text(header[col_idx])
                    cell_ref = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                    raw_value = self._cell_to_text(value)
                    chunks.append(Chunk(
                        doc_id=self.doc_id,
                        chunk_id=f"{self.doc_id}#{sheet_name}#{cell_ref}",
                        text=self._build_text(
                            sheet_name, cell_ref, row_label, column_header,
                            raw_value, unit, period, section_path,
                        ),
                        chunk_type="table_row",
                        source_title=self.source_title,
                        issuer=self.issuer,
                        doc_no="",
                        publish_date=self.publish_date,
                        section_path=section_path,
                        source_url=self.source_url,
                        local_path=self.local_path,
                        table_name=sheet_name,
                        indicator=row_label,
                        period=period,
                        unit=unit,
                        row_index=row_idx + 1,
                        cell_ref=cell_ref,
                        row_label=row_label,
                        column_header=column_header,
                        raw_value=raw_value,
                    ))

            for row_idx in range(header_idx + 1, end_idx):
                if row_idx in section_rows:
                    continue
                row = rows[row_idx]
                if any(
                    self._looks_like_data_cell(row[col_idx])
                    for col_idx in data_columns
                ):
                    continue
                texts = [
                    (col_idx, self._cell_to_text(value))
                    for col_idx, value in enumerate(row)
                    if not self._is_empty(value)
                ]
                if not texts or self._is_structural_text_row([text for _, text in texts]):
                    continue
                col_idx, text_val = texts[0]
                cell_ref = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                chunks.append(Chunk(
                    doc_id=self.doc_id,
                    chunk_id=f"{self.doc_id}#{sheet_name}#{cell_ref}",
                    text=self._build_text(
                        sheet_name, cell_ref, text_val, "", text_val,
                        unit, period, base_section,
                    ),
                    chunk_type="table_row",
                    source_title=self.source_title,
                    issuer=self.issuer,
                    doc_no="",
                    publish_date=self.publish_date,
                    section_path=base_section,
                    source_url=self.source_url,
                    local_path=self.local_path,
                    table_name=sheet_name,
                    indicator=text_val,
                    period=period,
                    unit=unit,
                    row_index=row_idx + 1,
                    cell_ref=cell_ref,
                    row_label=text_val,
                    column_header="",
                    raw_value=text_val,
                ))
        return chunks

    def _normalize_rows(self, rows: list[list[Any]]) -> list[list[Any]]:
        max_len = max((len(row) for row in rows), default=0)
        return [list(row) + [None] * (max_len - len(row)) for row in rows]

    def _detect_header_row(self, rows: list[list[Any]]) -> int | None:
        labels = {"项目", "指标", "指标名称", "地区", "metric", "item"}
        for row_idx, row in enumerate(rows):
            non_empty = [self._cell_to_text(value) for value in row if not self._is_empty(value)]
            if len(non_empty) < 2:
                continue
            if any(value in labels for value in non_empty):
                return row_idx
            if any(self._row_below_has_numeric(rows, row_idx, col_idx) for col_idx, _ in enumerate(row)):
                return row_idx
        return None

    def _detect_header_rows(self, rows: list[list[Any]]) -> list[int]:
        labels = {
            "项目", "指标", "指标名称", "地区", "机构",
            "metric", "item",
        }
        explicit = []
        for row_idx, row in enumerate(rows):
            non_empty = [
                self._cell_to_text(value)
                for value in row if not self._is_empty(value)
            ]
            if len(non_empty) < 2 or not any(value in labels for value in non_empty):
                continue
            if any(
                self._row_below_has_numeric(rows, row_idx, col_idx)
                for col_idx in range(len(row))
            ):
                explicit.append(row_idx)
        if explicit:
            return explicit
        fallback = self._detect_header_row(rows)
        return [fallback] if fallback is not None else []

    def _detect_data_columns(self, header: list[Any], data_rows: list[list[Any]]) -> list[int]:
        return [
            col_idx for col_idx, value in enumerate(header)
            if not self._is_empty(value)
            and any(
                col_idx < len(row) and self._looks_like_data_cell(row[col_idx])
                for row in data_rows
            )
        ]

    def _section_before_header(self, rows: list[list[Any]], header_idx: int,
                               lower_bound: int, sheet_name: str) -> tuple[str, int | None]:
        normalized_source = self._normalize_text(self.source_title)
        normalized_sheet = self._normalize_text(sheet_name)
        for row_idx in range(header_idx - 1, lower_bound - 1, -1):
            texts = [
                self._cell_to_text(value)
                for value in rows[row_idx] if not self._is_empty(value)
            ]
            if len(texts) != 1:
                continue
            text = texts[0]
            normalized = self._normalize_text(text)
            if not normalized or text.startswith(("单位", "unit", "Unit")):
                continue
            if text in {"时间", "项目", "指标"}:
                continue
            if normalized in normalized_source or normalized == normalized_sheet:
                continue
            return text, row_idx
        return "", None

    def _is_structural_text_row(self, texts: list[str]) -> bool:
        if any(text.startswith(("单位", "unit", "Unit")) for text in texts):
            return True
        return texts[0] in {"时间", "项目", "指标", "时间/指标"}

    def _dedupe_texts(self, values) -> list[str]:
        result = []
        for value in values:
            if value and value not in result:
                result.append(value)
        return result

    def _normalize_text(self, value: str) -> str:
        return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", str(value).lower())

    def _row_below_has_numeric(self, rows: list[list[Any]], header_idx: int, col_idx: int) -> bool:
        for row in rows[header_idx + 1: header_idx + 6]:
            if col_idx < len(row) and self._looks_like_data_cell(row[col_idx]):
                return True
        return False

    def _detect_label_column(self, header: list[Any]) -> int:
        labels = {"项目", "指标", "指标名称", "地区", "metric", "item"}
        for idx, value in enumerate(header):
            if self._cell_to_text(value) in labels:
                return idx
        for idx, value in enumerate(header):
            if not self._is_empty(value):
                return idx
        return 0

    def _detect_unit(self, rows: list[list[Any]]) -> str:
        for row in rows[:8]:
            for value in row:
                text = self._cell_to_text(value)
                if "单位" in text or text.lower().startswith("unit"):
                    return text
        return ""

    def _detect_period(self, texts: list[str]) -> str:
        joined = " ".join(texts)
        for pattern in [
            r"20\d{2}年\d{1,2}月",
            r"20\d{2}年[一二三四1-4]季度",
            r"20\d{2}Q[1-4]",
            r"20\d{2}[-/]\d{1,2}",
            r"20\d{2}",
        ]:
            match = re.search(pattern, joined)
            if match:
                return match.group(0)
        return ""

    def _build_text(self, sheet_name: str, cell_ref: str, row_label: str,
                    column_header: str, raw_value: str, unit: str, period: str,
                    section_path: list[str] = None) -> str:
        formatted_value = self._format_value(raw_value, row_label, unit)
        normalized_unit = self._normalize_unit(unit)
        parts = [
            f"文件《{self.source_title}》",
            f"工作表「{sheet_name}」",
            f"单元格 {cell_ref}",
        ]
        if section_path:
            parts.append(f"所属区块「{' > '.join(section_path)}」")
        parts.extend([
            f"行指标「{row_label}」",
            f"列口径「{column_header}」",
            f"原始值为 {formatted_value}",
        ])
        if normalized_unit:
            parts.append(f"单位：{normalized_unit}")
        if period:
            parts.append(f"期间：{period}")
        return "；".join(parts) + "。"

    def _normalize_unit(self, unit: str) -> str:
        if not unit:
            return ""
        text = unit.strip()
        while text.startswith("单位：") or text.startswith("单位:"):
            text = text.split("：", 1)[1] if "：" in text else text.split(":", 1)[1]
            text = text.strip()
        return text

    def _format_value(self, raw_value: str, row_label: str, unit: str) -> str:
        try:
            num = float(raw_value)
        except (ValueError, TypeError):
            return raw_value
        if self._looks_like_percent_row(row_label, unit) and abs(num) <= 1:
            return f"{num * 100:.2f}%"
        if num == int(num):
            return str(int(num))
        return f"{num:.2f}"

    def _looks_like_percent_row(self, row_label: str, unit: str) -> bool:
        label = row_label or ""
        unit_text = unit or ""
        return "%" in unit_text and any(
            kw in label for kw in ("增长率", "占比", "比例", "率", "比上年")
        )

    def _looks_like_data_cell(self, value: Any) -> bool:
        if isinstance(value, (int, float)):
            return True
        return bool(re.fullmatch(r"-?\d+(\.\d+)?%?", self._cell_to_text(value)))

    def _is_note_row(self, text: str) -> bool:
        return text.startswith(("注", "备注", "说明", "Notes", "Note"))

    def _has_any_value(self, row: list[Any]) -> bool:
        return any(not self._is_empty(value) for value in row)

    def _is_empty(self, value: Any) -> bool:
        return self._cell_to_text(value) == ""

    def _cell_to_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()
