#!/usr/bin/env python
"""运行100题专项评测，支持逐题落盘、断点续跑和错误题重试。"""
from __future__ import annotations

import argparse
from datetime import datetime
import json
import math
import os
from pathlib import Path
import re
import shutil
import statistics
import sys
import unicodedata

from dotenv import load_dotenv
from openpyxl import load_workbook


SOURCE_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(SOURCE_ROOT))
load_dotenv(SOURCE_ROOT / ".env")
os.environ.setdefault("HF_HUB_OFFLINE", "1")

from src.generator.answer_builder import AnswerBuilder
from src.generator.decomposer import QueryDecomposer
import src.generator.llm_client as llm_client_module


SPECIALIZED_SYSTEM_PROMPT = """你是银行保险监管资料库的开放式问答助手。只能依据【参考资料】作答。

规则：
1. 不得使用外部知识，不得用相近年份、相近机构或相近指标替代目标值。
2. 数值、日期、机构名称、文号必须与参考资料一致；计算题列出操作数、算式、单位和结果。
3. 跨文件问题必须分别引用支持每个操作数或判断条件的证据。
4. 若资料库完全没有所需资料，或只覆盖部分必要事实，将 answer 留空，behavior="refuse"，在 refuse_reason 中明确缺少什么。
5. 若问题缺少年份、期间、机构、指标或统计口径而存在多个合理解释，不得猜测；在 answer 中提出需要用户补充的具体条件，behavior="clarify"。
6. 只有资料足以支持唯一答案时 behavior="answer"。行业汇总值不能证明每家机构均符合要求，跨机构类型比较必须说明口径限制。
7. 严格输出 JSON，不要输出其他内容：
{
  "behavior": "answer|refuse|clarify",
  "answer": "回答或澄清问题；拒答时为空字符串",
  "evidence": [{"source_title":"文件名称","section":"位置","text":"证据原文","source_url":""}],
  "refuse_reason": null
}
"""


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).lower()
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", text)


def extract_numbers(text: str) -> list[float]:
    values = []
    normalized = unicodedata.normalize("NFKC", text or "").replace("−", "-").replace("–", "-")
    for match in re.finditer(r"(?<![\d.])[-+]?\d[\d,]*(?:\.\d+)?", normalized):
        try:
            values.append(float(match.group().replace(",", "")))
        except ValueError:
            continue
    return values


def chinese_integer(value: int) -> str | None:
    """Return the common Chinese form for a non-negative integer below 100."""
    digits = "零一二三四五六七八九"
    if not 0 <= value < 100:
        return None
    if value < 10:
        return digits[value]
    tens, ones = divmod(value, 10)
    prefix = "十" if tens == 1 else digits[tens] + "十"
    return prefix if ones == 0 else prefix + digits[ones]


def chinese_numeric_entity_match(entity: dict, text: str) -> bool:
    """Match exact small integer + unit forms such as 三年 and 一个函证基准日."""
    try:
        expected = float(entity["value"])
        tolerance = float(entity.get("tolerance") or 0)
    except (KeyError, TypeError, ValueError):
        return False
    if tolerance != 0 or not expected.is_integer():
        return False
    unit = str(entity.get("unit") or "").strip()
    chinese = chinese_integer(int(expected))
    if not unit or chinese is None:
        return False
    candidate = normalize(chinese + unit)
    return bool(candidate and candidate in normalize(text))


def entity_match(entity: dict, text: str) -> bool:
    if entity.get("match") == "numeric":
        expected = float(entity["value"])
        tolerance = float(entity.get("tolerance") or 0)
        scoped_text = text
        context_patterns = entity.get("context_patterns") or []
        if context_patterns:
            clauses = re.split(r"[。；;\n]", unicodedata.normalize("NFKC", text or ""))
            scoped_clauses = [
                clause for clause in clauses
                if any(re.search(pattern, clause, flags=re.IGNORECASE) for pattern in context_patterns)
            ]
            if not scoped_clauses:
                return False
            scoped_text = "\n".join(scoped_clauses)
        arabic_match = any(abs(actual - expected) <= tolerance + 1e-9 for actual in extract_numbers(scoped_text))
        return arabic_match or chinese_numeric_entity_match(entity, scoped_text)
    candidates = [str(entity.get("value", "")), *[str(alias) for alias in entity.get("aliases", [])]]
    normalized_text = normalize(text)
    return any(normalize(candidate) and normalize(candidate) in normalized_text for candidate in candidates)


def source_match(expected: str, actual: str) -> bool:
    a = normalize(expected).replace("财产保险公司", "财产险公司").replace("人身保险公司", "人身险公司")
    b = normalize(actual).replace("财产保险公司", "财产险公司").replace("人身保险公司", "人身险公司")
    return bool(a and b and (a in b or b in a))


def _contains_cell_ref(text: str, cell_ref: str) -> bool:
    return bool(re.search(rf"(?<![A-Z0-9]){re.escape(cell_ref)}(?!\d)", unicodedata.normalize("NFKC", text), re.IGNORECASE))


def evidence_item_match(requirement: dict, evidence: dict) -> bool:
    """核对单条证据的位置、期间、指标、数值或制度关键文本。"""
    combined = "\n".join(str(evidence.get(key) or "") for key in ("source_title", "section", "text"))
    if requirement.get("cell_ref") and not _contains_cell_ref(combined, str(requirement["cell_ref"])):
        return False
    if requirement.get("period") and normalize(requirement["period"]) not in normalize(combined):
        return False
    if requirement.get("column_header") and normalize(requirement["column_header"]) not in normalize(combined):
        return False
    if requirement.get("indicator") and normalize(requirement["indicator"]) not in normalize(combined):
        return False
    if "value" in requirement:
        expected = float(requirement["value"])
        tolerance = float(requirement.get("tolerance") or 0)
        if not any(abs(actual - expected) <= tolerance + 1e-9 for actual in extract_numbers(combined)):
            return False
    if any(normalize(term) not in normalize(combined) for term in requirement.get("text_terms", [])):
        return False
    return True


def exact_source_results(item: dict, evidence: list[dict]) -> tuple[list[dict], bool | None, str]:
    """跨文件题按结构化约束核验；没有约束时明确返回 unavailable。"""
    requirements = item.get("expected_evidence") or []
    if not requirements:
        return [], None, "unavailable"
    rows = []
    for requirement in requirements:
        expected_title = requirement["source_title"]
        candidates = [
            ev for ev in evidence
            if isinstance(ev, dict) and source_match(expected_title, ev.get("source_title", ""))
        ]
        item_results = []
        for expected_item in requirement.get("items", []):
            matched = any(evidence_item_match(expected_item, ev) for ev in candidates)
            item_results.append({"requirement": expected_item, "matched": matched})
        matched = bool(item_results) and all(row["matched"] for row in item_results)
        rows.append({"expected": expected_title, "matched": matched, "item_results": item_results})
    return rows, all(row["matched"] for row in rows), "available"


def classify_behavior(result: dict) -> str:
    explicit = str(result.get("behavior") or "").strip().lower()
    answer = str(result.get("answer") or "")
    refusal = str(result.get("refuse_reason") or "")
    clarify_terms = ("请补充", "请明确", "请提供", "需要明确", "需要补充", "请说明", "哪一年", "哪个季度", "何种口径")
    if explicit in {"answer", "refuse", "clarify"}:
        return explicit
    if any(term in answer for term in clarify_terms):
        return "clarify"
    if refusal or not answer.strip():
        return "refuse"
    return "answer"


def literal_forbidden_hits(item: dict, answer_text: str) -> list[str]:
    placeholders = {"具体数值", "无证据的结果值", "未经确认的具体答案"}
    normalized_answer = normalize(answer_text)
    return [
        value for value in item.get("forbidden_values", [])
        if value not in placeholders and normalize(value) and normalize(value) in normalized_answer
    ]


def clarification_dimension_covered(entity: dict, text: str) -> bool:
    """允许澄清答案使用等价维度词，不要求逐字复述标注。"""
    target = normalize(entity.get("value"))
    actual = normalize(text)
    dimension_groups = [
        (("年份", "季度", "时间", "时点", "期间", "月份", "日期"), ("年", "季度", "时间", "时点", "期间", "月", "日期")),
        (("机构", "公司", "银行", "范围", "主体", "类型"), ("机构", "公司", "银行", "范围", "主体", "类型")),
        (("指标", "口径", "分子", "分母", "数字", "单位"), ("指标", "口径", "分子", "分母", "数字", "数值", "单位", "单元格")),
        (("监管要求", "制度", "条款", "事项"), ("监管要求", "制度", "条款", "事项", "报告", "报送")),
    ]
    return any(any(key in target for key in targets) and any(key in actual for key in actuals) for targets, actuals in dimension_groups)


def score_result(item: dict, result: dict) -> dict:
    actual_answer = str(result.get("answer") or "")
    refusal = str(result.get("refuse_reason") or "")
    combined = "\n".join(part for part in (actual_answer, refusal) if part)
    behavior = classify_behavior(result)
    entity_results = [
        {"entity": entity, "matched": entity_match(entity, combined)}
        for entity in item.get("critical_entities", [])
    ]
    forbidden_hits = literal_forbidden_hits(item, actual_answer)
    evidence = result.get("evidence") if isinstance(result.get("evidence"), list) else []
    if item["category"] == "跨文件":
        source_results, evidence_complete, evidence_status = exact_source_results(item, evidence)
    else:
        source_results = []
        for expected in item.get("expected_sources", []):
            matched = any(source_match(expected, ev.get("source_title", "")) for ev in evidence if isinstance(ev, dict))
            source_results.append({"expected": expected, "matched": matched})
        evidence_complete = None
        evidence_status = "not_applicable"
    expected_behavior = item["expected_behavior"]
    failures = []
    if behavior != expected_behavior_map(expected_behavior):
        failures.append(f"预期行为={expected_behavior}，实际行为={behavior}")
    matched_count = sum(row["matched"] for row in entity_results)
    if item["category"] == "库外处理" and expected_behavior == "澄清":
        semantic_match = any(clarification_dimension_covered(row["entity"], combined) for row in entity_results)
        if matched_count == 0 and not semantic_match:
            failures.append("澄清未覆盖任何标注的核心缺失维度")
    elif expected_behavior == "回答":
        missing = [str(row["entity"].get("value")) for row in entity_results if not row["matched"]]
        if missing:
            failures.append("关键实体未命中：" + "、".join(missing))
    if forbidden_hits:
        failures.append("出现禁止错误值：" + "、".join(forbidden_hits))

    answer_correct = not failures
    is_correct = answer_correct
    if item["category"] == "跨文件" and evidence_status == "unavailable":
        is_correct = False
        failures.append("跨文件证据判定 unavailable：评测集缺少结构化证据约束")
    elif item["category"] == "跨文件" and not evidence_complete:
        is_correct = False
        failures.append("跨文件证据未完整命中：" + "、".join(row["expected"] for row in source_results if not row["matched"]))
    evaluation_status = item.get("evaluation_status", "valid")
    return {
        "actual_behavior": behavior,
        "entity_results": entity_results,
        "forbidden_hits": forbidden_hits,
        "source_results": source_results,
        "evidence_complete": evidence_complete,
        "evidence_status": evidence_status,
        "answer_correct": answer_correct,
        "is_correct": is_correct,
        "evaluation_status": evaluation_status,
        "included_in_primary": evaluation_status == "valid",
        "evaluation_note": item.get("evaluation_note", ""),
        "failure_reason": "；".join(failures),
    }


def expected_behavior_map(value: str) -> str:
    return {"回答": "answer", "拒答": "refuse", "澄清": "clarify"}[value]


def percentile(values: list[int], p: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    index = (len(ordered) - 1) * p
    low = math.floor(index)
    high = math.ceil(index)
    if low == high:
        return float(ordered[low])
    return ordered[low] * (high - index) + ordered[high] * (index - low)


def metric(correct: int, total: int, status: str = "available") -> dict:
    return {"status": status, "correct": correct, "total": total, "rate": round(correct / total, 4) if total else None}


def summarize(items: list[dict], results: list[dict]) -> dict:
    by_id = {item["id"]: item for item in items}
    successful = [result for result in results if not result.get("error")]
    errors = [result for result in results if result.get("error")]

    valid_results = [result for result in successful if result.get("scoring", {}).get("included_in_primary", True)]
    key_results = [result for result in valid_results if by_id[result["id"]]["category"] == "关键实体"]
    key_entities = [entity for result in key_results for entity in result["scoring"]["entity_results"]]
    entity_errors = sum(not row["matched"] for row in key_entities)

    refusal_results = [result for result in valid_results if by_id[result["id"]]["expected_behavior"] == "拒答"]
    clarification_results = [result for result in valid_results if by_id[result["id"]]["expected_behavior"] == "澄清"]
    cross_results = [result for result in valid_results if by_id[result["id"]]["category"] == "跨文件"]
    source_rows = [source for result in cross_results for source in result["scoring"]["source_results"]]
    latencies = [int(result["latency_ms"]) for result in successful if result.get("latency_ms") is not None]

    category_metrics = {}
    subtype_metrics = {}
    for item in items:
        result = next((row for row in results if row["id"] == item["id"]), None)
        if result is None:
            continue
        for target, key in ((category_metrics, item["category"]), (subtype_metrics, item["subtype"])):
            bucket = target.setdefault(key, {"correct": 0, "total": 0, "valid_correct": 0, "valid_total": 0})
            bucket["total"] += 1
            bucket["correct"] += int(bool(result.get("scoring", {}).get("is_correct")))
            if result.get("scoring", {}).get("included_in_primary", True):
                bucket["valid_total"] += 1
                bucket["valid_correct"] += int(bool(result.get("scoring", {}).get("is_correct")))
    for grouped in (category_metrics, subtype_metrics):
        for bucket in grouped.values():
            bucket["rate"] = round(bucket["correct"] / bucket["total"], 4) if bucket["total"] else None
            bucket["valid_rate"] = round(bucket["valid_correct"] / bucket["valid_total"], 4) if bucket["valid_total"] else None

    overall_correct = sum(bool(result.get("scoring", {}).get("is_correct")) for result in results)
    valid_correct = sum(bool(result.get("scoring", {}).get("is_correct")) for result in valid_results)
    ambiguous_ids = [result["id"] for result in successful if not result.get("scoring", {}).get("included_in_primary", True)]
    evidence_available = all(result["scoring"].get("evidence_status") == "available" for result in cross_results)
    return {
        "overall": metric(overall_correct, len(items)),
        "overall_valid": metric(valid_correct, len(valid_results)),
        "ambiguous_questions": {"count": len(ambiguous_ids), "ids": ambiguous_ids},
        "critical_entity_error_rate": {
            "status": "available" if key_entities else "unavailable",
            "errors": entity_errors,
            "total_entities": len(key_entities),
            "rate": round(entity_errors / len(key_entities), 4) if key_entities else None,
            "definition": "排除歧义题后的有效关键实体题中，未按文本规则或数值容差命中的金标准实体数 ÷ 有效金标准实体总数",
        },
        "out_of_scope": {
            "refusal_accuracy": metric(sum(r["scoring"]["actual_behavior"] == "refuse" and r["scoring"]["is_correct"] for r in refusal_results), len(refusal_results)),
            "clarification_accuracy": metric(sum(r["scoring"]["actual_behavior"] == "clarify" and r["scoring"]["is_correct"] for r in clarification_results), len(clarification_results)),
        },
        "cross_file_answer_accuracy": metric(sum(r["scoring"]["answer_correct"] for r in cross_results), len(cross_results)),
        "cross_file_evidence_hit_rate": metric(
            sum(row["matched"] for row in source_rows),
            len(source_rows),
            "available" if evidence_available else "unavailable",
        ),
        "cross_file_question_evidence_complete_rate": metric(
            sum(r["scoring"].get("evidence_complete") is True for r in cross_results),
            len(cross_results),
            "available" if evidence_available else "unavailable",
        ),
        "by_category": category_metrics,
        "by_subtype": subtype_metrics,
        "latency_ms": {
            "count": len(latencies),
            "average": round(statistics.mean(latencies), 1) if latencies else None,
            "median": round(statistics.median(latencies), 1) if latencies else None,
            "p95": round(percentile(latencies, .95), 1) if latencies else None,
            "min": min(latencies) if latencies else None,
            "max": max(latencies) if latencies else None,
        },
        "execution_errors": {"count": len(errors), "ids": [row["id"] for row in errors]},
    }


def _split_lines(value: object) -> list[str]:
    return [line.strip() for line in str(value or "").splitlines() if line.strip()]


def _json_lines(value: object, field: str, item_id: str) -> list[dict]:
    rows = []
    for line in _split_lines(value):
        try:
            parsed = json.loads(line)
        except json.JSONDecodeError as exc:
            raise RuntimeError(f"{item_id} 的{field}不是合法JSON：{exc}") from exc
        if not isinstance(parsed, dict):
            raise RuntimeError(f"{item_id} 的{field}每行必须是JSON对象")
        rows.append(parsed)
    return rows


def load_dataset(path: Path) -> list[dict]:
    """直接从专项Excel读取题目；Excel是唯一评测集交付文件。"""
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "专项评测集" not in workbook.sheetnames:
        raise RuntimeError(f"评测集缺少工作表“专项评测集”：{path}")
    sheet = workbook["专项评测集"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    required_headers = {
        "题号", "来源", "题型", "子类型", "问题", "标准答案", "预期行为", "关键实体",
        "预期来源文件", "预期证据约束", "评测有效性", "评测说明", "禁止出现的错误值", "评分规则",
    }
    missing_headers = sorted(required_headers - set(headers))
    if missing_headers:
        raise RuntimeError("评测集缺少列：" + "、".join(missing_headers))
    items = []
    for values in rows:
        row = dict(zip(headers, values))
        item_id = str(row.get("题号") or "").strip()
        if not item_id:
            continue
        items.append({
            "id": item_id,
            "origin": str(row.get("来源") or ""),
            "category": str(row.get("题型") or ""),
            "subtype": str(row.get("子类型") or ""),
            "question": str(row.get("问题") or ""),
            "standard_answer": str(row.get("标准答案") or ""),
            "expected_behavior": str(row.get("预期行为") or ""),
            "critical_entities": _json_lines(row.get("关键实体"), "关键实体", item_id),
            "expected_sources": _split_lines(row.get("预期来源文件")),
            "expected_evidence": _json_lines(row.get("预期证据约束"), "预期证据约束", item_id),
            "evaluation_status": str(row.get("评测有效性") or "valid"),
            "evaluation_note": str(row.get("评测说明") or ""),
            "forbidden_values": _split_lines(row.get("禁止出现的错误值")),
            "scoring_rule": str(row.get("评分规则") or ""),
        })
    expected_ids = [f"S{i:03d}" for i in range(1, 101)]
    if [item["id"] for item in items] != expected_ids:
        raise RuntimeError("专项评测集必须包含唯一连续的S001-S100")
    counts = {
        "关键实体": sum(item["category"] == "关键实体" for item in items),
        "库外处理": sum(item["category"] == "库外处理" for item in items),
        "跨文件": sum(item["category"] == "跨文件" for item in items),
    }
    if counts != {"关键实体": 50, "库外处理": 30, "跨文件": 20}:
        raise RuntimeError(f"专项评测集题型数量错误：{counts}")
    return items


def load_progress(path: Path) -> dict[str, dict]:
    recorded = {}
    if not path.exists():
        return recorded
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.strip():
            row = json.loads(line)
            recorded[row["id"]] = row
    return recorded


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dataset", type=Path, default=SOURCE_ROOT / "data/eval/银行监管RAG专项评测集_100题.xlsx")
    parser.add_argument("--run-dir", type=Path, help="默认写入 data/eval/runs/<run-name>/")
    parser.add_argument("--run-name", default="specialized-100-v1")
    parser.add_argument("--ids", default="", help="可选，逗号分隔题号")
    parser.add_argument("--request-timeout", type=float, default=60.0, help="专项评测单次LLM请求超时秒数")
    parser.add_argument("--publish", action="store_true", help="完整100题结束后复制为 data/eval/specialized_eval_report.json")
    parser.add_argument("--validate-only", action="store_true", help="只校验Excel结构，不调用模型")
    args = parser.parse_args()
    dataset_path = args.dataset if args.dataset.is_absolute() else SOURCE_ROOT / args.dataset
    items = load_dataset(dataset_path)
    if args.validate_only:
        print(json.dumps({"dataset": str(dataset_path), "total": len(items), "status": "valid"}, ensure_ascii=False, indent=2))
        return
    if args.ids:
        selected = [value.strip().upper() for value in args.ids.split(",") if value.strip()]
        by_id = {item["id"]: item for item in items}
        missing = [item_id for item_id in selected if item_id not in by_id]
        if missing:
            parser.error("不存在题号：" + ", ".join(missing))
        items = [by_id[item_id] for item_id in selected]
    if args.publish and len(items) != 100:
        parser.error("--publish 只允许完整100题运行，不能与 --ids 局部运行同时使用")

    run_dir = args.run_dir or SOURCE_ROOT / "data/eval/runs" / args.run_name
    if not run_dir.is_absolute():
        run_dir = SOURCE_ROOT / run_dir
    run_dir.mkdir(parents=True, exist_ok=True)
    progress_path = run_dir / "progress.jsonl"
    report_path = run_dir / "report.json"
    recorded = load_progress(progress_path)
    done_ids = {item_id for item_id, row in recorded.items() if not row.get("error")}
    if done_ids:
        print(f"[续跑] 已有 {len(done_ids)} 题成功记录；仅重试错误或未运行题。", flush=True)
    llm_client_module._MAX_ATTEMPTS = 1
    builder = AnswerBuilder(decomposer=QueryDecomposer(include_single_fact_options=False))
    builder.llm.client = builder.llm.client.with_options(timeout=args.request_timeout, max_retries=0)
    builder.decomposer.llm.client = builder.decomposer.llm.client.with_options(timeout=args.request_timeout, max_retries=0)
    started_at = datetime.now().astimezone().isoformat(timespec="seconds")
    with progress_path.open("a", encoding="utf-8") as progress:
        for index, item in enumerate(items, 1):
            if item["id"] in done_ids:
                continue
            print(f"[{index}/{len(items)}] {item['id']} {item['category']}/{item['subtype']}", flush=True)
            try:
                result = builder.answer(item["question"], system_prompt=SPECIALIZED_SYSTEM_PROMPT, include_diagnostics=True)
                scoring = score_result(item, result)
                diagnostics = result.get("diagnostics", {})
                record = {
                    **{key: item[key] for key in ("id", "origin", "category", "subtype", "question", "standard_answer", "expected_behavior", "critical_entities", "expected_sources", "expected_evidence", "forbidden_values", "scoring_rule", "evaluation_status", "evaluation_note")},
                    "actual_answer": result.get("answer", ""),
                    "refuse_reason": result.get("refuse_reason"),
                    "raw_behavior": result.get("behavior"),
                    "evidence": result.get("evidence", []),
                    "latency_ms": result.get("latency_ms"),
                    "timing_ms": diagnostics.get("timing_ms", {}),
                    "llm_metrics": diagnostics.get("llm", {}),
                    "retrieval": diagnostics.get("retrieval", {}),
                    "routing": diagnostics.get("routing", {}),
                    "scoring": scoring,
                }
            except Exception as exc:
                record = {
                    **{key: item[key] for key in ("id", "origin", "category", "subtype", "question", "standard_answer", "expected_behavior", "critical_entities", "expected_sources", "expected_evidence", "forbidden_values", "scoring_rule", "evaluation_status", "evaluation_note")},
                    "actual_answer": "", "refuse_reason": None, "evidence": [], "latency_ms": None,
                    "scoring": {"is_correct": False, "answer_correct": False, "evidence_complete": False, "failure_reason": str(exc)},
                    "error": f"{type(exc).__name__}: {exc}",
                }
                print(f"  [执行错误] {record['error']}", flush=True)
            recorded[item["id"]] = record
            progress.write(json.dumps(record, ensure_ascii=False) + "\n")
            progress.flush()

    results = [recorded[item["id"]] for item in items if item["id"] in recorded]
    metrics = summarize(items, results)
    report = {
        "run_name": args.run_name,
        "dataset": str(dataset_path),
        "model": os.environ.get("LLM_MODEL", "gpt-4o-mini"),
        "started_at": started_at,
        "completed_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "selected_ids": [item["id"] for item in items],
        "total": metrics["overall"]["total"],
        "correct": metrics["overall"]["correct"],
        "accuracy": metrics["overall"]["rate"],
        "metrics": metrics,
        "results": results,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    published_path = None
    if args.publish:
        published_path = SOURCE_ROOT / "data/eval/specialized_eval_report.json"
        shutil.copyfile(report_path, published_path)
    print(json.dumps({
        "report": str(report_path),
        "published_report": str(published_path) if published_path else None,
        "metrics": report["metrics"],
    }, ensure_ascii=False, indent=2), flush=True)


if __name__ == "__main__":
    main()
