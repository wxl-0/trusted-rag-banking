"""
验证 QA 测评集的 evidence 能否在已解析的 chunk 中找到。

用法:
    python scripts/check_qa_evidence_coverage.py
    python scripts/check_qa_evidence_coverage.py --source-type word
    python scripts/check_qa_evidence_coverage.py --verbose

输出:
    按 source_type 分组的覆盖率，以及未命中样例。
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
QA_PATH = REPO_ROOT / "data" / "eval" / "QA数据.xlsx"
CLAUSE_PATH = REPO_ROOT / "data" / "chunks" / "clause_chunks.jsonl"
TABLE_PATH = REPO_ROOT / "data" / "chunks" / "table_chunks.jsonl"


def load_chunks() -> tuple[list[str], list[str]]:
    """Load all chunk texts and raw_values from JSONL files."""
    texts = []
    raw_values = []
    for path in (CLAUSE_PATH, TABLE_PATH):
        if not path.exists():
            continue
        with open(path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                obj = json.loads(line)
                texts.append(obj.get("text", ""))
                rv = obj.get("raw_value")
                if rv is not None:
                    raw_values.append(str(rv))
    return texts, raw_values


def normalize_number(s: str) -> str:
    """Remove thousand separators for number matching."""
    return re.sub(r"(?<=\d),(?=\d{3})", "", s)


def _parse_excel_evidence(evidence: str) -> dict:
    """Parse structured excel evidence.

    Supports two formats:
      Format A: '...工作表：X；单元格：Y；单位：Z；原始值：V。'
      Format B: '...指标=值(单元格)；指标2=值2(单元格2)；...'
    """
    result = {}
    m = re.search(r"工作表[：:]\s*(.+?)(?=[；;]|$)", evidence)
    if m:
        result["sheet"] = m.group(1).strip()
    m = re.search(r"单元格[：:]\s*([A-Z]+\d+)", evidence, re.I)
    if m:
        result["cell_ref"] = m.group(1).upper()
    m = re.search(r"原始值[：:]\s*([^。；]+)", evidence)
    if m:
        result["raw_value"] = m.group(1).strip()

    # Format B: extract all values from "指标=值(cell)" patterns
    if not result.get("raw_value"):
        values = re.findall(r"=(-?\d+(?:\.\d+)?)\(", evidence)
        if values:
            result["raw_values_list"] = values
    return result


def _strip_for_match(s: str) -> str:
    """Remove whitespace and normalize punctuation for fuzzy substring matching."""
    s = re.sub(r"\s+", "", s)
    s = s.replace("，", ",").replace("；", ";").replace("：", ":").replace("。", ".")
    s = s.replace("、", ",").replace("（", "(").replace("）", ")")
    return s


def evidence_matches(evidence: str, source_type: str, texts: list[str], raw_values: list[str],
                     stripped_texts: list[str] | None = None) -> bool:
    """Check if evidence can be found in chunks."""
    if not evidence or len(evidence.strip()) < 3:
        return False

    if source_type == "excel":
        parsed = _parse_excel_evidence(evidence)
        rv = parsed.get("raw_value", "")
        values_list = parsed.get("raw_values_list", [])

        if rv:
            if rv in raw_values:
                return True
            formatted = _add_thousand_sep(rv) if re.match(r"^-?\d+\.?\d*$", rv) else rv
            for t in texts:
                if rv in t or formatted in t:
                    return True

        # Format B: check if ANY of the extracted values exist in raw_values or texts
        if values_list:
            for v in values_list:
                if v in raw_values:
                    return True
                # Also check integer form (e.g. "34878.78" might be stored as "34878")
                try:
                    num = float(v)
                    int_form = str(int(num)) if num == int(num) else None
                    if int_form and int_form in raw_values:
                        return True
                except (ValueError, TypeError):
                    pass
            # If at least one value found in any chunk text
            for v in values_list[:3]:
                for t in texts:
                    if v in t:
                        return True
        return False
    else:
        # Word/PDF: evidence is actual text snippet
        snippets = [s.strip() for s in evidence.split("；") if len(s.strip()) >= 8]
        if not snippets:
            snippets = [evidence.strip()[:50]]

        # Strategy 1: sliding window on raw text (fast path)
        for snippet in snippets[:3]:
            if len(snippet) < 10:
                continue
            positions = [0, 5, 10, 15, 20, len(snippet) // 2]
            for start in positions:
                key = snippet[start:start + 15]
                if len(key) < 10:
                    continue
                for t in texts:
                    if key in t:
                        return True

        # Strategy 2: normalized match (handles whitespace/punctuation differences)
        if stripped_texts is not None:
            for snippet in snippets[:3]:
                if len(snippet) < 8:
                    continue
                stripped_snippet = _strip_for_match(snippet)
                if len(stripped_snippet) < 8:
                    continue
                positions = list(range(0, min(len(stripped_snippet) - 7, 30), 3))
                positions.append(len(stripped_snippet) // 2)
                for start in positions:
                    key = stripped_snippet[start:start + 8]
                    if len(key) < 8:
                        continue
                    for st in stripped_texts:
                        if key in st:
                            return True
        return False


def _add_thousand_sep(num_str: str) -> str:
    """Add thousand separators: '31739.18' -> '31,739.18'"""
    if "." in num_str:
        int_part, dec_part = num_str.split(".", 1)
    else:
        int_part, dec_part = num_str, None

    negative = int_part.startswith("-")
    if negative:
        int_part = int_part[1:]

    result = ""
    for i, ch in enumerate(reversed(int_part)):
        if i > 0 and i % 3 == 0:
            result = "," + result
        result = ch + result

    if negative:
        result = "-" + result
    if dec_part is not None:
        result = result + "." + dec_part
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="Check QA evidence coverage against chunks.")
    parser.add_argument("--qa-path", default=str(QA_PATH))
    parser.add_argument("--source-type", choices=["excel", "word", "pdf"], default=None)
    parser.add_argument("--verbose", action="store_true", help="Show all misses")
    parser.add_argument("--miss-limit", type=int, default=5, help="Max miss examples per type")
    args = parser.parse_args()

    qa_path = Path(args.qa_path)
    if not qa_path.exists():
        print(f"ERROR: QA file not found: {qa_path}")
        return 1

    if not CLAUSE_PATH.exists() and not TABLE_PATH.exists():
        print("ERROR: No chunk files found. Run scripts/ingest.py first.")
        return 1

    print("Loading chunks...", flush=True)
    texts, raw_values = load_chunks()
    print(f"  Loaded {len(texts)} chunks, {len(raw_values)} raw_values")

    print("Building normalized index...", flush=True)
    stripped_texts = [_strip_for_match(t) for t in texts]

    print("Loading QA dataset...", flush=True)
    from openpyxl import load_workbook
    wb = load_workbook(qa_path, read_only=True, data_only=True)
    ws = wb.active

    # Results tracking
    results: dict[str, dict] = {}
    misses: dict[str, list] = {}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"  Loaded {len(rows)} QA questions")

    for row in rows:
        qid = str(row[0] or "")
        source_type = str(row[1] or "").strip()
        evidence = row[12]  # col 13 (0-indexed col 12) = evidence
        source_title = str(row[13] or "")

        if args.source_type and source_type != args.source_type:
            continue

        if source_type not in results:
            results[source_type] = {"hit": 0, "miss": 0}
            misses[source_type] = []

        evidence_str = str(evidence) if evidence is not None else ""

        if evidence_matches(evidence_str, source_type, texts, raw_values, stripped_texts):
            results[source_type]["hit"] += 1
        else:
            results[source_type]["miss"] += 1
            limit = 999 if args.verbose else args.miss_limit
            if len(misses[source_type]) < limit:
                misses[source_type].append({
                    "qid": qid,
                    "evidence": evidence_str[:80],
                    "source_title": source_title[:50],
                })

    wb.close()

    # Print results
    print("\n" + "=" * 60)
    print("QA Evidence Coverage Report")
    print("=" * 60)

    total_hit = 0
    total_all = 0
    for stype in sorted(results.keys()):
        hit = results[stype]["hit"]
        miss = results[stype]["miss"]
        total = hit + miss
        pct = hit / total * 100 if total else 0
        total_hit += hit
        total_all += total
        print(f"\n  {stype:8s}: {hit:3d}/{total:3d} ({pct:5.1f}%)")

        if misses[stype]:
            print(f"    Misses (showing {len(misses[stype])}):")
            for m in misses[stype]:
                print(f"      {m['qid']}: ev=\"{m['evidence']}\"")
                print(f"              src=\"{m['source_title']}\"")

    overall_pct = total_hit / total_all * 100 if total_all else 0
    print(f"\n  {'TOTAL':8s}: {total_hit:3d}/{total_all:3d} ({overall_pct:5.1f}%)")
    print("=" * 60)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
