#!/usr/bin/env python
"""
对评测集跑问答，输出 eval_report.json。
用法：python scripts/run_eval.py [--limit N] [--source excel|word|pdf]
     python scripts/run_eval.py --ids Q035,Q068 --run-name diverse-baseline-v1
"""
import argparse
from datetime import datetime
import json
import os
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

sys.path.insert(0, str(Path(__file__).parent.parent))

load_dotenv()

# 嵌入/重排模型已在本地缓存，必须在导入生成模块前启用离线模式。
os.environ.setdefault("HF_HUB_OFFLINE", "1")

from src.generator.answer_builder import AnswerBuilder
from src.generator.decomposer import QueryDecomposer

QA_PATH = Path("data/eval/QA数据.xlsx")
REPORT_PATH = Path("data/eval/eval_report.json")
PROGRESS_PATH = Path("data/eval/eval_progress.jsonl")
RUNS_DIR = Path("data/eval/runs")

REGULATION_FACT_QA_TYPES = {"单事实检索", "多事实检索"}
TABLE_QA_TYPES = {"表格取数", "表格比较", "表格计算"}

EVAL_SYSTEM_PROMPT = """你是银行业监管制度选择题问答助手。请严格依据下方提供的参考资料回答问题。

规则：
1. 只能使用参考资料中的内容，禁止引入外部知识
2. 题目会给出 A、B、C、D 四个选项，必须在 choice 字段明确选择一个选项
3. 比较或计算题先在 answer 中列出资料数值和计算过程，再给出结论
4. 无法依据资料作答时，choice 返回 null，并在 refuse_reason 中说明原因
5. 严格按照 JSON 格式输出，不要输出其他内容

输出格式（JSON）：
{
  "choice": "A",
  "answer": "答案解释和依据",
  "evidence": [
    {
      "source_title": "文件名称",
      "section": "章节位置",
      "text": "原文片段",
      "source_url": "来源URL"
    }
  ],
  "refuse_reason": null
}"""


def load_qa_items(source_filter=None, item_ids=None, qa_path=QA_PATH):
    """从 QA数据.xlsx 加载选择题，返回列表。"""
    wb = openpyxl.load_workbook(qa_path)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    h = {v: i for i, v in enumerate(headers)}

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        source_type = row[h["source_type"]]
        if source_filter and source_type != source_filter:
            continue
        eval_status = (
            row[h["eval_status"]]
            if "eval_status" in h and h["eval_status"] < len(row)
            else ""
        )
        items.append(
            {
                "id": row[h["id"]],
                "source_type": source_type,
                "difficulty": row[h["difficulty"]],
                "qa_type": row[h["qa_type"]],
                "question": row[h["question"]],
                "option_a": row[h["option_a"]],
                "option_b": row[h["option_b"]],
                "option_c": row[h["option_c"]],
                "option_d": row[h["option_d"]],
                "answer": row[h["answer"]],          # 正确选项字母 A/B/C/D
                "answer_text": row[h["answer_text"]], # 正确答案的具体内容
                "evidence": row[h["evidence"]] if "evidence" in h else "",
                "source_title": row[h["source_title"]],
                "eval_status": eval_status or "",
            }
        )
    if item_ids is None:
        return [
            item for item in items
            if not item["eval_status"].startswith("excluded")
        ]

    items_by_id = {item["id"]: item for item in items}
    missing_ids = [item_id for item_id in item_ids if item_id not in items_by_id]
    if missing_ids:
        raise ValueError(f"评测集不存在题号：{', '.join(missing_ids)}")
    return [items_by_id[item_id] for item_id in item_ids]


def build_mc_question(item):
    """将选择题拼成完整提问文本（含选项）。"""
    return (
        f"{item['question']}\n"
        f"A. {item['option_a']}\n"
        f"B. {item['option_b']}\n"
        f"C. {item['option_c']}\n"
        f"D. {item['option_d']}"
    )


def extract_choice(answer_text: str) -> str:
    """从 RAG 回答中提取选项字母（A/B/C/D）。"""
    if not answer_text:
        return ""
    patterns = (
        r"^\s*([ABCD])(?:[\.、。\)）:：\s]|$)",
        r"(?:答案|选项|选择|choice)\s*(?:是|为|[:：])?\s*([ABCD])",
        r"([ABCD])\s*选项",
    )
    for pattern in patterns:
        match = re.search(pattern, answer_text, flags=re.IGNORECASE)
        if match:
            return match.group(1).upper()
    return ""


def score_answer(result: dict, item: dict) -> dict:
    """使用可复现规则把模型输出映射到选择题选项。"""
    if result.get("refuse_reason"):
        return {
            "choice": "",
            "is_correct": False,
            "scoring_method": "refused",
        }

    structured_choice = str(result.get("choice") or "").strip().upper()
    if structured_choice in {"A", "B", "C", "D"}:
        return {
            "choice": structured_choice,
            "is_correct": structured_choice == item["answer"],
            "scoring_method": "structured_choice",
        }

    answer_text = str(result.get("answer") or "")
    if not answer_text.strip():
        return {
            "choice": "",
            "is_correct": False,
            "scoring_method": "empty_answer",
        }

    explicit_choice = extract_choice(answer_text)
    if explicit_choice:
        return {
            "choice": explicit_choice,
            "is_correct": explicit_choice == item["answer"],
            "scoring_method": "explicit_choice_text",
        }

    normalized_answer = _normalize_for_match(answer_text)
    matching_choices = []
    for choice in ("A", "B", "C", "D"):
        normalized_option = _normalize_for_match(item[f"option_{choice.lower()}"])
        if len(normalized_option) >= 2 and normalized_option in normalized_answer:
            matching_choices.append(choice)
    if len(matching_choices) == 1:
        choice = matching_choices[0]
        return {
            "choice": choice,
            "is_correct": choice == item["answer"],
            "scoring_method": "normalized_option_text",
        }

    return {
        "choice": "",
        "is_correct": False,
        "scoring_method": "unparseable",
    }


def _normalize_for_match(value) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).lower()
    return re.sub(r"[\s,，。；;：:、]", "", text)


def parse_item_ids(value: str) -> list[str]:
    item_ids = [item_id.strip().upper() for item_id in value.split(",") if item_id.strip()]
    if not item_ids:
        raise argparse.ArgumentTypeError("--ids 至少需要一个题号")
    if len(item_ids) != len(set(item_ids)):
        raise argparse.ArgumentTypeError("--ids 中不能包含重复题号")
    return item_ids


def resolve_output_paths(run_name: str | None) -> tuple[Path, Path]:
    if not run_name:
        return REPORT_PATH, PROGRESS_PATH
    if run_name in {".", ".."} or not re.fullmatch(r"[A-Za-z0-9._-]+", run_name):
        raise ValueError("--run-name 只能包含字母、数字、点、下划线和连字符")
    run_dir = RUNS_DIR / run_name
    return run_dir / "report.json", run_dir / "progress.jsonl"


def summarize_accuracy(results: list[dict], field: str) -> dict:
    """按指定题目字段汇总数量、正确数和准确率。"""
    grouped: dict[str, dict] = {}
    for result in results:
        key = result.get(field) or "unknown"
        bucket = grouped.setdefault(key, {"total": 0, "correct": 0})
        bucket["total"] += 1
        bucket["correct"] += int(bool(result.get("is_correct")))
    return {
        key: {
            "total": values["total"],
            "correct": values["correct"],
            "accuracy": round(values["correct"] / values["total"], 4),
        }
        for key, values in grouped.items()
    }


def normalize_evidence_title(title: str) -> str:
    normalized = re.sub(
        r"[\W_]+",
        "",
        unicodedata.normalize("NFKC", str(title or "")).lower(),
    )
    normalized = normalized.replace("财产保险公司", "财产险公司")
    normalized = normalized.replace("人身保险公司", "人身险公司")
    return normalized


def build_competition_metrics(results: list[dict]) -> dict:
    """按赛题建议口径聚合可复现的量化指标。"""
    regulation_results = [
        result
        for result in results
        if result.get("qa_type") in REGULATION_FACT_QA_TYPES
    ]
    total = len(regulation_results)
    correct = sum(bool(result.get("is_correct")) for result in regulation_results)
    rate = round(correct / total, 4) if total else None
    table_results = [
        result
        for result in results
        if result.get("qa_type") in TABLE_QA_TYPES
    ]
    table_total = len(table_results)
    table_correct = sum(bool(result.get("is_correct")) for result in table_results)
    table_rate = round(table_correct / table_total, 4) if table_total else None
    evidence_results = [
        result for result in results if result.get("expected_source_title")
    ]
    evidence_hits = 0
    for result in evidence_results:
        expected_title = normalize_evidence_title(result["expected_source_title"])
        actual_titles = {
            normalize_evidence_title(evidence.get("source_title") or "")
            for evidence in (result.get("evidence") or [])
        }
        evidence_hits += int(
            any(expected_title in actual_title for actual_title in actual_titles)
        )
    evidence_total = len(evidence_results)
    evidence_rate = (
        round(evidence_hits / evidence_total, 4) if evidence_total else None
    )
    refused_count = sum(bool(result.get("refused")) for result in results)
    observed_refusal_rate = (
        round(refused_count / len(results), 4) if results else None
    )
    return {
        "regulation_fact_accuracy": {
            "status": "available" if total else "unavailable",
            "total": total,
            "correct": correct,
            "rate": rate,
            "target": 0.85,
            "meets_target": rate >= 0.85 if rate is not None else None,
        },
        "table_accuracy": {
            "status": "available" if table_total else "unavailable",
            "total": table_total,
            "correct": table_correct,
            "rate": table_rate,
            "target": 0.8,
            "meets_target": table_rate >= 0.8 if table_rate is not None else None,
        },
        "evidence_citation_hit_rate": {
            "status": "available" if evidence_total else "unavailable",
            "evaluated": evidence_total,
            "hits": evidence_hits,
            "rate": evidence_rate,
            "target": 0.9,
            "meets_target": (
                evidence_rate >= 0.9 if evidence_rate is not None else None
            ),
            "definition": (
                "返回证据中至少一个 source_title 经规范化及标题别名处理后"
                "包含题库标准来源标题"
            ),
        },
        "critical_entity_error_rate": {
            "status": "unavailable",
            "rate": None,
            "target_max": 0.05,
            "meets_target": None,
            "reason": "题库尚无结构化的数字、日期、机构名称和文号金标准标注",
        },
        "out_of_scope_refusal_rate": {
            "status": "unavailable",
            "rate": None,
            "target": 0.8,
            "meets_target": None,
            "observed_refusal_count": refused_count,
            "observed_refusal_rate": observed_refusal_rate,
            "reason": "当前评测集均为可回答选择题，尚无库外或依据不足题目标注",
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=None, help="只跑前 N 题，用于调试")
    parser.add_argument("--source", choices=["excel", "word", "pdf"], default=None)
    parser.add_argument("--ids", type=parse_item_ids, default=None,
                        help="按给定顺序运行题号，使用英文逗号分隔")
    parser.add_argument("--run-name", default=None,
                        help="将进度和报告保存到 data/eval/runs/<run-name>/")
    args = parser.parse_args()

    if args.ids and (args.limit is not None or args.source is not None):
        parser.error("--ids 不能与 --limit 或 --source 同时使用")
    try:
        report_path, progress_path = resolve_output_paths(args.run_name)
        items = load_qa_items(source_filter=args.source, item_ids=args.ids)
    except ValueError as exc:
        parser.error(str(exc))
    if args.limit:
        items = items[: args.limit]

    report_path.parent.mkdir(parents=True, exist_ok=True)
    builder = AnswerBuilder(
        decomposer=QueryDecomposer(include_single_fact_options=True)
    )
    started_at = datetime.now().astimezone().isoformat(timespec="seconds")
    if args.run_name:
        print(f"[运行] {args.run_name}；题目：{', '.join(item['id'] for item in items)}")

    # 断点续传：加载已有进度（同一 id 取最后一次记录）
    recorded: dict = {}
    if progress_path.exists():
        for line in progress_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line:
                continue
            r = json.loads(line)
            recorded[r["id"]] = r
    done_ids = {rid for rid, r in recorded.items() if not r.get("error")}
    if done_ids:
        print(f"[续传] 已完成 {len(done_ids)} 题，跳过；出错的题将重试")

    with progress_path.open("a", encoding="utf-8") as progress_fp:
        for item in items:
            if item["id"] in done_ids:
                continue
            full_q = build_mc_question(item)
            print(f"[{item['id']}] {item['source_type']}/{item['difficulty']} {item['question'][:40]}...")
            options = {
                "A": item["option_a"],
                "B": item["option_b"],
                "C": item["option_c"],
                "D": item["option_d"],
            }

            try:
                result = builder.answer(
                    full_q,
                    system_prompt=EVAL_SYSTEM_PROMPT,
                    include_diagnostics=True,
                )
                actual = result.get("answer", "")
                refuse = result.get("refuse_reason")
                score = score_answer(result, item)
                diagnostics = result.get("diagnostics", {})
                record = {
                    "id": item["id"],
                    "source_type": item["source_type"],
                    "difficulty": item["difficulty"],
                    "qa_type": item.get("qa_type", "unknown"),
                    "question": item["question"],
                    "options": options,
                    "correct_answer": item["answer"],
                    "correct_answer_text": item["answer_text"],
                    "expected_evidence": item.get("evidence", ""),
                    "expected_source_title": item.get("source_title", ""),
                    "raw_choice": result.get("choice"),
                    "choice": score["choice"],
                    "scoring_method": score["scoring_method"],
                    "actual_answer": actual,
                    "evidence": result.get("evidence", []),
                    "is_correct": score["is_correct"],
                    "refused": bool(refuse),
                    "refuse_reason": refuse,
                    "latency_ms": result.get("latency_ms"),
                    "timing_ms": diagnostics.get("timing_ms", {}),
                    "llm_metrics": diagnostics.get("llm", {}),
                    "sub_questions": diagnostics.get("sub_questions", []),
                    "routing": diagnostics.get("routing", {}),
                    "retrieval": diagnostics.get("retrieval", {}),
                }
            except Exception as e:  # 单题异常不拖垮整轮，记为错误后继续
                print(f"    [出错] {e}")
                record = {
                    "id": item["id"],
                    "source_type": item["source_type"],
                    "difficulty": item["difficulty"],
                    "qa_type": item.get("qa_type", "unknown"),
                    "question": item["question"],
                    "options": options,
                    "correct_answer": item["answer"],
                    "correct_answer_text": item["answer_text"],
                    "expected_evidence": item.get("evidence", ""),
                    "expected_source_title": item.get("source_title", ""),
                    "raw_choice": None,
                    "choice": "",
                    "scoring_method": "error",
                    "actual_answer": "",
                    "evidence": [],
                    "is_correct": False,
                    "refused": False,
                    "refuse_reason": None,
                    "latency_ms": None,
                    "timing_ms": {},
                    "llm_metrics": {},
                    "sub_questions": [],
                    "routing": {},
                    "retrieval": {},
                    "error": str(e),
                }

            recorded[item["id"]] = record
            progress_fp.write(json.dumps(record, ensure_ascii=False) + "\n")
            progress_fp.flush()

    # 从记录聚合本轮范围内的结果
    results = [recorded[item["id"]] for item in items if item["id"] in recorded]
    items_by_id = {item["id"]: item for item in items}
    for result in results:
        item = items_by_id[result["id"]]
        result["expected_evidence"] = item.get("evidence", "")
        result["expected_source_title"] = item.get("source_title", "")
    total = len(results)
    correct = sum(1 for r in results if r["is_correct"])
    errored = [r["id"] for r in results if r.get("error")]
    unparseable = [r["id"] for r in results if r.get("scoring_method") == "unparseable"]

    source_summary = summarize_accuracy(results, "source_type")
    qa_type_summary = summarize_accuracy(results, "qa_type")
    difficulty_summary = summarize_accuracy(results, "difficulty")
    scoring_methods: dict[str, int] = {}
    for result in results:
        method = result.get("scoring_method", "legacy")
        scoring_methods[method] = scoring_methods.get(method, 0) + 1

    timing_summary = {}
    for stage in ("decomposition", "retrieval", "generation", "total"):
        values = [
            result.get("timing_ms", {}).get(stage)
            for result in results
            if result.get("timing_ms", {}).get(stage) is not None
        ]
        if values:
            timing_summary[stage] = {
                "average_ms": round(sum(values) / len(values), 1),
                "min_ms": min(values),
                "max_ms": max(values),
            }

    total_api_calls = sum(
        result.get("llm_metrics", {}).get("total_api_calls", 0) or 0
        for result in results
    )
    token_values = [
        result.get("llm_metrics", {}).get("total_tokens")
        for result in results
        if result.get("llm_metrics", {}).get("total_tokens") is not None
    ]
    cost_values = [
        result.get("llm_metrics", {}).get("provider_reported_cost")
        for result in results
        if result.get("llm_metrics", {}).get("provider_reported_cost") is not None
    ]
    retrieval_targets = [
        target
        for result in results
        for target in result.get("retrieval", {}).get("targets", [])
    ]
    covered_targets = sum(bool(target.get("covered")) for target in retrieval_targets)
    coverage_statuses: dict[str, int] = {}
    for target in retrieval_targets:
        status = target.get("coverage_status") or (
            "supported" if target.get("covered") else "missing"
        )
        coverage_statuses[status] = coverage_statuses.get(status, 0) + 1
    supplemental_searches = sum(
        result.get("retrieval", {}).get("supplemental_searches", 0) or 0
        for result in results
    )
    report = {
        "run_name": args.run_name,
        "started_at": started_at,
        "completed_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "model": os.environ.get("LLM_MODEL", "gpt-4o-mini"),
        "selected_ids": [item["id"] for item in items],
        "total": total,
        "correct": correct,
        "accuracy": round(correct / total, 4) if total else 0,
        "scoring_methods": scoring_methods,
        "unparseable_ids": unparseable,
        "measurements": {
            "timing": timing_summary,
            "total_api_calls": total_api_calls,
            "total_tokens": sum(token_values) if token_values else None,
            "provider_reported_cost": sum(cost_values) if cost_values else None,
            "retrieval": {
                "target_count": len(retrieval_targets),
                "covered_target_count": covered_targets,
                "coverage_statuses": coverage_statuses,
                "coverage_rate": (
                    round(covered_targets / len(retrieval_targets), 4)
                    if retrieval_targets else None
                ),
                "supplemental_searches": supplemental_searches,
            },
        },
        "competition_metrics": build_competition_metrics(results),
        "by_source": source_summary,
        "by_qa_type": qa_type_summary,
        "by_difficulty": difficulty_summary,
        "results": results,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n评测完成：{correct}/{total}，准确率 {report['accuracy']:.1%}")
    for src, v in source_summary.items():
        print(f"  {src}: {v['correct']}/{v['total']} = {v['accuracy']:.1%}")
    print(f"  --- 按题型 ---")
    for qt, v in qa_type_summary.items():
        print(f"  {qt}: {v['correct']}/{v['total']} = {v['accuracy']:.1%}")
    print(f"  --- 按难度 ---")
    for difficulty, v in difficulty_summary.items():
        print(
            f"  {difficulty}: {v['correct']}/{v['total']} = "
            f"{v['accuracy']:.1%}"
        )
    if errored:
        print(f"  [注意] {len(errored)} 题执行出错（已记为错误）：{errored}")
        print(f"         重跑本脚本会自动重试这些题；如需从头跑请先删除 {progress_path}")
    if unparseable:
        print(f"  [人工复核] 无法确定选项的题目：{unparseable}")
    print(f"报告保存至：{report_path}")


if __name__ == "__main__":
    main()
